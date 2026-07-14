"""Reconstruct financial tables into Excel workbooks (Category-based sheets)."""

from __future__ import annotations

import io
from pathlib import Path
from typing import Any

try:
    import openpyxl  # type: ignore[import-not-found]
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ===================================================================
# Style constants
# ===================================================================

_FONT_TITLE = Font(name="Calibri", size=16, bold=True)
_FONT_SUBTITLE = Font(name="Calibri", size=14, bold=True, italic=True)
_FONT_SECTION = Font(name="Calibri", size=12, bold=True, color="1F4E79")
_FONT_SUBSECTION = Font(name="Calibri", size=10, bold=True, color="2E75B6")
_FONT_HEADER = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
_FONT_DATA = Font(name="Calibri", size=10)
_FONT_TOTAL = Font(name="Calibri", size=10, bold=True)

_FILL_HEADER = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_FILL_SECTION = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_FILL_TOTAL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_FILL_ALT_ROW = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
_ALIGN_TOP_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)

_BORDER_THIN = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

# ===================================================================
# Main builder
# ===================================================================

def build_excel(extraction_result: dict[str, Any]) -> bytes:
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required for Excel export.")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    metadata = extraction_result.get("metadata", {})
    master_sections = extraction_result.get("master_sections", [])
    text_extractions = extraction_result.get("text_extractions", [])
    table_extractions = extraction_result.get("table_extractions", [])

    _build_metadata_sheet(wb, extraction_result)
    
    if master_sections:
        _build_master_sections_sheet(wb, master_sections)
    if extraction_result.get("table_inventory"):
        _build_table_inventory_sheet(wb, extraction_result["table_inventory"])

    # Build Category Sheets
    categories = {}
    for sec in master_sections:
        cat = sec.get("category", "Uncategorized")
        if cat not in categories:
            categories[cat] = []
        categories[cat].append(sec)

    for cat, sections in categories.items():
        _build_category_sheet(wb, cat, sections, text_extractions, table_extractions)

    # Build Intelligence Report sheet
    structured_intel = extraction_result.get("structured_intelligence", {})
    if structured_intel:
        try:
            from .workbook_population import populate_intelligence_report
            report_rows = populate_intelligence_report(structured_intel, master_sections)
            _build_intelligence_sheet(wb, report_rows)
        except Exception as exc:
            import logging
            logging.getLogger(__name__).warning(f"Failed to build intelligence sheet: {exc}")

    # Move Intelligence Report to first position (after Metadata)
    if "Intelligence Report" in wb.sheetnames:
        wb.move_sheet("Intelligence Report", offset=-len(wb.sheetnames) + 2)

    # Move Metadata sheet to first position
    if "Metadata" in wb.sheetnames:
        wb.move_sheet("Metadata", offset=-len(wb.sheetnames) + 1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def save_excel(extraction_result: dict[str, Any], path: str | Path) -> Path:
    path = Path(path)
    path.write_bytes(build_excel(extraction_result))
    return path


# ===================================================================
# Intelligence Sheet Builder
# ===================================================================

def _build_intelligence_sheet(wb: "openpyxl.Workbook", rows: list[dict]) -> None:
    ws = wb.create_sheet(title="Intelligence Report")
    
    # Headers
    headers = ["Category", "Sub Category", "Status", "Extracted Value", "Source Page", "Confidence"]
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = _FONT_HEADER
        c.fill = _FILL_HEADER
        c.alignment = _ALIGN_CENTER
        c.border = _BORDER_THIN
        
    for row_idx, r in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=r["category"]).alignment = _ALIGN_TOP_LEFT
        ws.cell(row=row_idx, column=2, value=r["subcategory"]).alignment = _ALIGN_TOP_LEFT
        ws.cell(row=row_idx, column=3, value=r.get("status", "")).alignment = _ALIGN_TOP_LEFT
        
        val_cell = ws.cell(row=row_idx, column=4, value=r["extracted_value"])
        val_cell.alignment = _ALIGN_LEFT  # wrap_text=True
        
        ws.cell(row=row_idx, column=5, value=r["source_page"]).alignment = _ALIGN_TOP_LEFT
        ws.cell(row=row_idx, column=6, value=r["confidence"]).alignment = _ALIGN_TOP_LEFT
        
        # Apply borders
        for col_idx in range(1, 7):
            ws.cell(row=row_idx, column=col_idx).border = _BORDER_THIN
            
    # Apply alt row fills
    for row_idx in range(2, len(rows) + 2, 2):
        for col_idx in range(1, 7):
            ws.cell(row=row_idx, column=col_idx).fill = _FILL_ALT_ROW
            
    # Widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 80
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15


# ===================================================================
# Category Sheet Builder
# ===================================================================

def _build_category_sheet(
    wb: "openpyxl.Workbook",
    category_name: str,
    sections: list[dict],
    text_extractions: list[dict],
    table_extractions: list[dict]
) -> None:
    # Ensure sheet name is valid
    invalid_chars = r"[]:*?/"
    safe_name = "".join(c for c in category_name if c not in invalid_chars)
    safe_name = safe_name[:31]
    ws = wb.create_sheet(title=safe_name)

    current_row = 1

    for section in sections:
        section_name = section.get("section_name", "Unknown Section")
        
        # Print Section Title
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
        cell = ws.cell(row=current_row, column=1, value=f"{section_name}")
        cell.font = _FONT_TITLE
        cell.fill = _FILL_SECTION
        cell.alignment = _ALIGN_LEFT
        current_row += 2

        if section.get("content_type") == "table":
            table_match = _find_table(section, table_extractions)
            if table_match:
                current_row = _render_table_block(ws, current_row, table_match)
            else:
                ws.cell(row=current_row, column=1, value="(Table data not extracted)").font = _FONT_DATA
                current_row += 2
        else:
            text_match = _find_text(section, text_extractions)
            if text_match:
                current_row = _render_text_block(ws, current_row, text_match)
            else:
                ws.cell(row=current_row, column=1, value="(Text data not extracted)").font = _FONT_DATA
                current_row += 2
                
        current_row += 2

    # Column widths
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20


def _find_table(section: dict, table_extractions: list[dict]) -> dict | None:
    sec_name = section.get("section_name", "")
    
    # 1. Check exact match for generic table
    for table in table_extractions:
        if table.get("table_name") == sec_name:
            return table
            
    # 2. Check legacy map
    LEGACY_MAP = {
        "Standalone Balance Sheet": "standalone_balance_sheet",
        "Standalone Profit & Loss": "standalone_profit_and_loss",
        "Standalone Cash Flow": "standalone_cash_flow",
        "Consolidated Balance Sheet": "consolidated_balance_sheet",
        "Consolidated Profit & Loss": "consolidated_profit_and_loss",
        "Consolidated Cash Flow": "consolidated_cash_flow",
    }
    
    if sec_name in LEGACY_MAP:
        mapped = LEGACY_MAP[sec_name]
        for table in table_extractions:
            if table.get("table_name") == mapped:
                return table
                
    return None

def _find_text(section: dict, text_extractions: list[dict]) -> str | None:
    cat = section.get("category")
    subcat = section.get("section_name")
    for ext in text_extractions:
        if ext.get("category") == cat and ext.get("subcategory") == subcat:
            return ext.get("extracted_text")
    return None


def _render_text_block(ws: "openpyxl.worksheet.worksheet.Worksheet", start_row: int, text: str) -> int:
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=10)
    cell = ws.cell(row=start_row, column=1, value=text)
    cell.font = _FONT_DATA
    cell.alignment = _ALIGN_TOP_LEFT
    
    # Estimate height (roughly 90 chars per line across 10 cols)
    lines = len(text.splitlines()) + (len(text) // 90)
    ws.row_dimensions[start_row].height = max(15, lines * 15)
    
    return start_row + 1

def _render_table_block(ws: "openpyxl.worksheet.worksheet.Worksheet", start_row: int, table_extract: dict) -> int:
    table_json = table_extract.get("table_json", {})
    if not table_json:
        ws.cell(row=start_row, column=1, value="No table data").font = _FONT_DATA
        return start_row + 1
        
    current_row = start_row

    title = table_json.get("title", "")
    if title:
        ws.cell(row=current_row, column=1, value=title).font = _FONT_SUBTITLE
        current_row += 1

    currency = table_json.get("currency", "")
    if currency:
        ws.cell(row=current_row, column=1, value=f"({currency})").font = _FONT_DATA
        current_row += 1

    # Legacy tables have dicts in rows, Generic tables have arrays
    rows = table_json.get("rows", [])
    if not rows:
        return current_row
        
    if isinstance(rows[0], list):
        # GENERIC TABLE
        headers = table_json.get("column_headers", [])
        if headers:
            for col_idx, h in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=h)
                cell.font = _FONT_HEADER
                cell.fill = _FILL_HEADER
                cell.alignment = _ALIGN_CENTER
                cell.border = _BORDER_THIN
            current_row += 1
            
        for row_data in rows:
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = _FONT_DATA
                cell.border = _BORDER_THIN
            current_row += 1
    else:
        # LEGACY FINANCIAL STATEMENT
        col_headers = table_json.get("column_headers", ["Note No.", "Current Period", "Previous Period"])
        has_note = any(h.lower().replace(".", "").replace(" ", "") == "noteno" for h in col_headers)
        headers = ["Particulars"] + col_headers

        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=h)
            cell.font = _FONT_HEADER
            cell.fill = _FILL_HEADER
            cell.alignment = _ALIGN_CENTER
            cell.border = _BORDER_THIN
        current_row += 1

        prev_section_parts = []
        for row_data in rows:
            section = row_data.get("section")
            line_item = row_data.get("line_item", "")
            note_no = row_data.get("note_no")
            values = row_data.get("values", {})
            current_val = values.get("current_period")
            previous_val = values.get("previous_period")

            is_total = "total" in line_item.lower()

            section_parts = section.split(" > ") if section else []
            for depth, part in enumerate(section_parts):
                if depth < len(prev_section_parts) and prev_section_parts[depth] == part:
                    continue
                for insert_depth in range(depth, len(section_parts)):
                    p = section_parts[insert_depth]
                    if not p:
                        continue
                    indent = "    " * insert_depth
                    cell = ws.cell(row=current_row, column=1, value=f"{indent}{p}")
                    cell.font = _FONT_SECTION if insert_depth == 0 else _FONT_SUBSECTION
                    cell.border = _BORDER_THIN
                    for c in range(2, len(headers) + 1):
                        ws.cell(row=current_row, column=c).border = _BORDER_THIN
                    current_row += 1
                break
            prev_section_parts = section_parts

            indent = "    " * len(section_parts)
            display_label = f"{indent}{line_item}" if line_item else ""

            cell = ws.cell(row=current_row, column=1, value=display_label)
            cell.font = _FONT_TOTAL if is_total else _FONT_DATA
            cell.border = _BORDER_THIN
            if is_total:
                cell.fill = _FILL_TOTAL

            col_offset = 2
            if has_note:
                cell_note = ws.cell(row=current_row, column=col_offset, value=note_no)
                cell_note.alignment = _ALIGN_CENTER
                cell_note.border = _BORDER_THIN
                col_offset += 1

            for val in (current_val, previous_val):
                cell_val = ws.cell(row=current_row, column=col_offset)
                if val is not None:
                    cell_val.value = val
                    cell_val.number_format = '#,##0.00;[Red](#,##0.00)'
                else:
                    cell_val.value = "-"
                    cell_val.alignment = _ALIGN_CENTER
                cell_val.font = _FONT_TOTAL if is_total else _FONT_DATA
                cell_val.border = _BORDER_THIN
                if is_total:
                    cell_val.fill = _FILL_TOTAL
                col_offset += 1

            current_row += 1

    return current_row


# ===================================================================
# Existing Base Sheets (Metadata, Master Sections, Inventory)
# ===================================================================

def _build_metadata_sheet(wb: "openpyxl.Workbook", extraction_result: dict[str, Any]) -> None:
    ws = wb.create_sheet(title="Metadata")
    _FONT_META_KEY = Font(name="Calibri", size=10, bold=True, color="1F4E79")
    _FONT_META_VAL = Font(name="Calibri", size=10)
    
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    cell = ws.cell(row=1, column=1, value="Extraction Metadata")
    cell.font = _FONT_TITLE
    
    row = 3
    metadata = extraction_result.get("metadata", {})
    for key, label in [
        ("file_name", "File Name"),
        ("financial_year", "Financial Year"),
        ("page_count", "Page Count"),
        ("extraction_timestamp", "Timestamp")
    ]:
        ws.cell(row=row, column=1, value=label).font = _FONT_META_KEY
        ws.cell(row=row, column=2, value=str(metadata.get(key, ""))).font = _FONT_META_VAL
        row += 1
        
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 50

def _build_master_sections_sheet(wb: "openpyxl.Workbook", master_sections: list[dict]) -> None:
    ws = wb.create_sheet(title="Master Sections")
    headers = ["Section ID", "Section Name", "Category", "Subcategory", "Start Page", "End Page",
               "Page Count", "Content Type", "Strategy", "Source", "Confidence"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
    
    for row_idx, section in enumerate(master_sections, start=2):
        ws.cell(row=row_idx, column=1, value=section.get("section_id"))
        ws.cell(row=row_idx, column=2, value=section.get("section_name"))
        ws.cell(row=row_idx, column=3, value=section.get("category"))
        ws.cell(row=row_idx, column=4, value=section.get("subcategory", ""))
        ws.cell(row=row_idx, column=5, value=section.get("start_page"))
        ws.cell(row=row_idx, column=6, value=section.get("end_page"))
        ws.cell(row=row_idx, column=7, value=section.get("page_count",
                    section.get("end_page", 0) - section.get("start_page", 0) + 1))
        ws.cell(row=row_idx, column=8, value=section.get("content_type"))
        ws.cell(row=row_idx, column=9, value=section.get("extraction_strategy"))
        ws.cell(row=row_idx, column=10, value=section.get("source", "taxonomy"))
        ws.cell(row=row_idx, column=11, value=section.get("confidence"))
        
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 25
    ws.freeze_panes = "A2"

def _build_table_inventory_sheet(wb: "openpyxl.Workbook", inventory: list[dict]) -> None:
    ws = wb.create_sheet(title="Table Inventory")
    headers = ["Table ID", "Table Name", "Table Category", "Page No.", "Needs VLM", "Parent Section"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        
    for row_idx, item in enumerate(inventory, start=2):
        ws.cell(row=row_idx, column=1, value=item.get("table_id"))
        ws.cell(row=row_idx, column=2, value=item.get("table_name"))
        ws.cell(row=row_idx, column=3, value=item.get("table_category", "other"))
        ws.cell(row=row_idx, column=4, value=item.get("page_no"))
        ws.cell(row=row_idx, column=5, value=str(item.get("needs_vlm", False)))
        ws.cell(row=row_idx, column=6, value=item.get("parent_section_id", ""))
        
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["F"].width = 25
    ws.freeze_panes = "A2"

